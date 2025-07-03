import pytest
import pandas as pd
from app.export import export_data
from openpyxl import load_workbook

@pytest.mark.parametrize('ext,fmt', [
    ('.csv', 'csv'),
    ('.xlsx', 'xlsx'),
    ('.json', 'json'),
    ('.txt', 'txt'),
    ('.html', 'html')
])
def test_export_formats(tmp_path, sample_df, fmt, ext):
    file_path = tmp_path / f'test_export{ext}'
    export_data(sample_df, file_path, fmt)
    assert file_path.exists(), f"Exported file {file_path} does not exist."

    # Check that the export file is valid and contains expected content
    content_samples = {'Lord-Captain Demos Carinus has retired from the service at the age of 41. Current Assignment: Unassigned', 'Commander Iacton Akios has retired from the service at the age of 32. Current Assignment: Unassigned', 'Commander Ravan Malephar awarded the Loyal Servant based on the condition: Ten Years of Service'}

    # Separate tests for excel and other formats
    if fmt == 'xlsx':
        wb = load_workbook(file_path).active

        data_rows = list(wb.iter_rows(min_row=2, values_only=True))
        assert any(any(s in str(cell) for s in content_samples) for row in data_rows for cell in row), f"Exported Excel file {file_path} does not contain expected content."
    else:
        content = file_path.read_text(encoding='utf-8') if ext != '.xlsx' else None
        assert all(s in content for s in content_samples), f"Exported file {file_path} does not contain expected content."

def test_export_excel_logging_failure(tmp_path, mock_logger_export, sample_df, mocker):
    mock_to_excel = mocker.patch.object(sample_df, 'to_excel', side_effect=Exception("Excel export failed"))
    file_path = tmp_path / 'test_export.xlsx'

    # Test the specific error handling for Excel export
    with pytest.raises(IOError, match="Failed to export to Excel"):
        export_data(sample_df, file_path, 'xlsx')

    mock_to_excel.assert_called_once_with(file_path, index=False)

def test_export_unsupported_format(tmp_path, sample_df):
    # Test that the app logs an error when trying to export using an unsupported format
    file_path = tmp_path / 'test_export.unsupported'
    with pytest.raises(ValueError, match="Unsupported format: unsupported"):
        export_data(sample_df, file_path, 'unsupported')

    assert not file_path.exists(), f"Exported file {file_path} should not exist for unsupported format."

def test_export_empty_dataframe(tmp_path, mock_logger_export):
    # Test that the app does not export an empty DataFrame and logs a warning
    file_path = tmp_path / 'empty_export.csv'
    empty_df = pd.DataFrame()
    export_data(empty_df, file_path, 'csv')

    assert not file_path.exists(), f"Exported file {file_path} does not exist for empty DataFrame."
    mock_logger_export.warning.assert_called_once
    assert "DataFrame is empty." in mock_logger_export.warning.call_args[0][0], "Expected warning about empty DataFrame not logged."

def test_logging_on_success(mock_logger_export, tmp_path, sample_df):
    # Tests app logging on successful export
    file_path = tmp_path / 'test_export.csv'
    export_data(sample_df, file_path, 'csv')

    # Check that the logger was called with the correct message
    mock_logger_export.info.assert_called_once_with(f"Data successfully exported to {file_path}")